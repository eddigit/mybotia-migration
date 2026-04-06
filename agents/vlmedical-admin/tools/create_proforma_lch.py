#!/usr/bin/env python3
"""
Génère la proforma VL Medical → Émergence SAS pour gants LCH Sensitex
Format identique à la proforma HEXA N° 260106 mais avec LCH au même prix
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()
ws = wb.active
ws.title = "Proforma"

# --- Page setup ---
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = 'portrait'
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# --- Largeurs colonnes ---
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 18

# --- Styles ---
title_font = Font(name='Arial', size=16, bold=True, color='003366')
subtitle_font = Font(name='Arial', size=9, italic=True, color='666666')
header_font = Font(name='Arial', size=10, bold=True, color='003366')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
small_font = Font(name='Arial', size=8)
small_italic = Font(name='Arial', size=8, italic=True)
money_font = Font(name='Arial', size=10)
total_font = Font(name='Arial', size=11, bold=True, color='003366')

header_fill = PatternFill(start_color='E8EEF5', end_color='E8EEF5', fill_type='solid')
total_fill = PatternFill(start_color='D4E1F0', end_color='D4E1F0', fill_type='solid')

thin_border = Border(
    bottom=Side(style='thin', color='CCCCCC')
)
bottom_border = Border(
    bottom=Side(style='medium', color='003366')
)

align_right = Alignment(horizontal='right', vertical='center')
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_center = Alignment(horizontal='center', vertical='center')

EUR_FMT = '#,##0.00 €'

row = 1

# === EN-TÊTE VL MEDICAL ===
ws.merge_cells('A1:E1')
ws['A1'] = 'VL MEDICAL'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

row = 2
ws.merge_cells('A2:E2')
ws['A2'] = 'Dispositifs médicaux, équipements de protection individuelle & trading divers'
ws['A2'].font = subtitle_font
ws['A2'].alignment = Alignment(horizontal='center')

row = 3
ws.merge_cells('A3:E3')
ws['A3'] = '853 225 100 R.C.S. Aix-en-Provence'
ws['A3'].font = small_font
ws['A3'].alignment = Alignment(horizontal='center')

row = 4
ws.merge_cells('A4:E4')
ws['A4'] = '190 Rue Topaze 13510 Éguilles'
ws['A4'].font = small_font
ws['A4'].alignment = Alignment(horizontal='center')

row = 5
ws.merge_cells('A5:E5')
ws['A5'] = 'Numéro EORI : FR85322510000025'
ws['A5'].font = small_font
ws['A5'].alignment = Alignment(horizontal='center')

row = 7
# Client info
ws.merge_cells('C7:E7')
ws['C7'] = 'Émergence Internationale SAS'
ws['C7'].font = bold_font

ws.merge_cells('C8:E8')
ws['C8'] = 'M. OUATTARA  af.emergence@gmail.com'
ws['C8'].font = normal_font

ws.merge_cells('C9:E9')
ws['C9'] = '9 bis passage Dartois Bidot,'
ws['C9'].font = normal_font

ws.merge_cells('C10:E10')
ws['C10'] = '94100 Saint-Maur-des-Fossés'
ws['C10'].font = normal_font

ws.merge_cells('C11:E11')
ws['C11'] = 'Siret : 89451323300011    TVA : FR82894513233'
ws['C11'].font = small_font

# Date
row = 13
ws['A13'] = 'ÉGUILLES, le lundi 10 mars 2026'
ws['A13'].font = bold_font

row = 15
ws.merge_cells('A15:E15')
ws['A15'] = 'PHARMACEUTICAL PRODUCTS — TRADE TERMS : CIF Cotonou'
ws['A15'].font = Font(name='Arial', size=11, bold=True, color='003366')
ws['A15'].alignment = Alignment(horizontal='center')

row = 17
ws['A17'] = 'Proforma N° 260301'
ws['A17'].font = Font(name='Arial', size=12, bold=True, color='003366')

# === TABLEAU PRODUITS ===
row = 19
headers = ['Libellé', 'Base', 'Master boxes', 'Price', 'Amount']
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = bottom_border

row = 20
ws.merge_cells('A20:E20')
ws['A20'] = 'Natural rubber gloves, non-sterile, powder-free - Master boxes of ten Boxes of 100 gloves'
ws['A20'].font = Font(name='Arial', size=9, italic=True)

# Lignes produit
products = [
    ('Latex LCH Sensitex Cartons de 10 Boîtes Size S', 2, 140, 15.50),
    ('Latex LCH Sensitex Cartons de 10 Boîtes Size M', 22, 1540, 15.50),
    ('Latex LCH Sensitex Cartons de 10 Boîtes Size L', 22, 1540, 15.50),
    ('Latex LCH Sensitex Cartons de 10 Boîtes Size XL', 2, 140, 15.50),
]

row = 21
for desc, base, boxes, price in products:
    ws.cell(row=row, column=1, value=desc).font = normal_font
    ws.cell(row=row, column=2, value=base).font = normal_font
    ws.cell(row=row, column=2).alignment = align_center
    ws.cell(row=row, column=3, value=boxes).font = normal_font
    ws.cell(row=row, column=3).alignment = align_center
    ws.cell(row=row, column=4, value=price).font = money_font
    ws.cell(row=row, column=4).number_format = EUR_FMT
    ws.cell(row=row, column=4).alignment = align_right
    amount = boxes * price
    ws.cell(row=row, column=5, value=amount).font = money_font
    ws.cell(row=row, column=5).number_format = EUR_FMT
    ws.cell(row=row, column=5).alignment = align_right
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = thin_border
    row += 1

# Avoir no show
ws.cell(row=row, column=1, value='Avoir en valeur sur no show précédente expédition').font = normal_font
ws.cell(row=row, column=5, value=-360.00).font = money_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
for c in range(1, 6):
    ws.cell(row=row, column=c).border = thin_border
row += 1

# Ligne vide
row += 1

# Info maritime
ws.cell(row=row, column=1, value='Compagnie maritime : MSC                POL: Fos-sur-Mer').font = small_font
row += 1
ws.cell(row=row, column=1, value='Transit time : 32 jours via Lomé').font = small_font
row += 1
row += 1

# --- Calculs ---
marchandise_total = (140 * 15.50) + (1540 * 15.50) + (1540 * 15.50) + (140 * 15.50) - 360.00
# = 2170 + 23870 + 23870 + 2170 - 360 = 51720

# Assurance — base = marchandise brute + fret + chargement - avoir + BESC (méthode proforma originale)
assurance_base_110 = 52080 + 3900 + 1450 - 360 + 85  # = 57155
assurance_taux = 0.0072
assurance = round(assurance_base_110 * assurance_taux, 2)  # = 411.52

# Insurance
ws.cell(row=row, column=1, value=f'Insurance base achats x 1,1 + transports').font = normal_font
ws.cell(row=row, column=2, value=assurance_base_110).font = small_font
ws.cell(row=row, column=2).number_format = EUR_FMT
ws.cell(row=row, column=3, value='0,72%').font = small_font
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=4, value=assurance).font = money_font
ws.cell(row=row, column=4).number_format = EUR_FMT
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=assurance).font = money_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
for c in range(1, 6):
    ws.cell(row=row, column=c).border = thin_border
row += 1

# BESC
ws.cell(row=row, column=1, value='BESC Cotonou si requis :').font = normal_font
ws.cell(row=row, column=3, value=1).font = normal_font
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=4, value=85.00).font = money_font
ws.cell(row=row, column=4).number_format = EUR_FMT
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=85.00).font = money_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
for c in range(1, 6):
    ws.cell(row=row, column=c).border = thin_border
row += 1

# Chargement
ws.cell(row=row, column=1, value='Chargement du Container de 40 pieds HQ').font = normal_font
ws.cell(row=row, column=3, value=1).font = normal_font
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=4, value=1450.00).font = money_font
ws.cell(row=row, column=4).number_format = EUR_FMT
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=1450.00).font = money_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
for c in range(1, 6):
    ws.cell(row=row, column=c).border = thin_border
row += 1

# Fret Centrimex
ws.cell(row=row, column=1, value='Cost of delivery to Cotonou (Centrimex offer)').font = normal_font
ws.cell(row=row, column=3, value=1).font = normal_font
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=4, value=3900.00).font = money_font
ws.cell(row=row, column=4).number_format = EUR_FMT
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=3900.00).font = money_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
for c in range(1, 6):
    ws.cell(row=row, column=c).border = bottom_border
row += 1

# Total master boxes
ws.cell(row=row, column=1, value='Estimation Chargement').font = bold_font
ws.cell(row=row, column=2, value='Total Master boxes').font = bold_font
ws.cell(row=row, column=3, value=3360).font = bold_font
ws.cell(row=row, column=3).alignment = align_center
row += 2

# === TOTAUX ===
total_ht = marchandise_total + assurance + 85 + 1450 + 3900

ws.cell(row=row, column=1, value='Incoterm:').font = bold_font
ws.cell(row=row, column=2, value='CIF').font = bold_font
ws.cell(row=row, column=4, value='Montant HT').font = total_font
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=total_ht).font = total_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
ws.cell(row=row, column=5).fill = total_fill
row += 1

ws.cell(row=row, column=1, value='Origin of goods:').font = bold_font
ws.cell(row=row, column=2, value='Malaysia').font = normal_font
ws.cell(row=row, column=4, value='TVA').font = normal_font
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value='No VAT export').font = normal_font
ws.cell(row=row, column=5).alignment = align_right
row += 1

ws.cell(row=row, column=1, value='HS code:').font = bold_font
ws.cell(row=row, column=2, value='40151900').font = normal_font
ws.merge_cells(f'C{row}:E{row}')
ws[f'C{row}'] = 'Marchandise destinée à l\'exportation et non soumise à TVA'
ws[f'C{row}'].font = small_italic
row += 1

ws.cell(row=row, column=1, value='Boxes Latex Gloves').font = bold_font
ws.cell(row=row, column=2, value=33600).font = bold_font
ws.merge_cells(f'C{row}:E{row}')
ws[f'C{row}'] = 'Marchandise remise au port de Fos-sur-Mer zone douanière'
ws[f'C{row}'].font = small_italic
row += 1

ws.cell(row=row, column=1, value='Containers').font = bold_font
ws.cell(row=row, column=2, value=1).font = bold_font
ws.cell(row=row, column=4, value='Net à Payer TTC').font = total_font
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=total_ht).font = total_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
ws.cell(row=row, column=5).fill = total_fill
row += 2

# Paiement
ws.cell(row=row, column=1, value='Condition de paiement : Comptant').font = bold_font
ws.cell(row=row, column=4, value='À la commande').font = normal_font
ws.cell(row=row, column=4).alignment = align_right
ws.cell(row=row, column=5, value=22000.00).font = bold_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
row += 1

ws.cell(row=row, column=1, value='Adresse de livraison :').font = normal_font
ws.cell(row=row, column=4, value='Au chargement').font = normal_font
ws.cell(row=row, column=4).alignment = align_right
solde = total_ht - 22000
ws.cell(row=row, column=5, value=solde).font = bold_font
ws.cell(row=row, column=5).number_format = EUR_FMT
ws.cell(row=row, column=5).alignment = align_right
row += 1

ws.cell(row=row, column=1, value='Port de Cotonou').font = normal_font
row += 1

ws.cell(row=row, column=1, value='Délai de mise à disposition des gants à Marseille auprès du transitaire ==> 2 jours après confirmation du devis et dépôt de l\'acompte').font = small_italic
ws.cell(row=row, column=1).alignment = align_left
row += 2

# Destination finale
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Marchandises exportées vers le Niger et livrées sous Douane. Client Final : Nord Sud Pharma SA'
ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True)
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Château 1, PL, rue des Dallols, BP. 2074 Niamey, Niger.'
ws[f'A{row}'].font = small_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 2

# Coordonnées bancaires
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Payment by bank transfer to our account at Société Générale'
ws[f'A{row}'].font = normal_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Paiement par virement sur notre compte Bancaire à la Société Générale'
ws[f'A{row}'].font = normal_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'SAS VL MEDICAL'
ws[f'A{row}'].font = bold_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = '190 Rue Topaze 13510 Éguilles'
ws[f'A{row}'].font = normal_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'IBAN FR76 3000 3001 6400 0200 0411 902'
ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'BIC : SOGEFRPP'
ws[f'A{row}'].font = bold_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 2

# Pied de page
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Société par action simplifiée de 1.000 €'
ws[f'A{row}'].font = small_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'N° TVA Intracommunautaire : FR 27853225100'
ws[f'A{row}'].font = small_font
ws[f'A{row}'].alignment = Alignment(horizontal='center')

# Print
print(f"Marchandise (gants) : {140*15.50 + 1540*15.50 + 1540*15.50 + 140*15.50:.2f} €")
print(f"Avoir no show : -360.00 €")
print(f"Sous-total marchandise : {marchandise_total:.2f} €")
print(f"Assurance base (x1.1) : {assurance_base_110:.2f} €")
print(f"Assurance (0.72%) : {assurance:.2f} €")
print(f"BESC : 85.00 €")
print(f"Chargement : 1 450.00 €")
print(f"Fret Centrimex : 3 900.00 €")
print(f"TOTAL HT = TTC : {total_ht:.2f} €")
print(f"Acompte : 22 000.00 €")
print(f"Solde : {solde:.2f} €")

output_path = '/home/gilles/.openclaw-vlmedical/workspace-vlmedical-admin/Proforma_VL_Emergence_LCH_260301.xlsx'
wb.save(output_path)
print(f"\nFichier sauvegardé : {output_path}")
